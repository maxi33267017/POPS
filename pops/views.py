from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
from .models import SoldEquipment, ServiceRecord
from decimal import Decimal
from dateutil.relativedelta import relativedelta
from django.db.models import Count, Sum, Min, Max
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from django.views.decorators.http import require_http_methods
from django.db.models.functions import TruncMonth
from django.db.models import Q
import io

def index(request):
    return render(request, 'pops/index.html')

@csrf_exempt
@require_http_methods(["POST"])
def upload_sold_equipment(request):
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': 'No se proporcionó ningún archivo'})

        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()

        if file_extension == 'json':
            data = json.loads(file.read().decode('utf-8'))
        elif file_extension in ['xlsx', 'xls']:
            # Leer el archivo Excel
            df = pd.read_excel(file)
            
            # Convertir el DataFrame a una lista de diccionarios
            data = df.to_dict('records')
            
            # Asegurarse de que las fechas estén en el formato correcto
            for record in data:
                if 'sale_date' in record:
                    if isinstance(record['sale_date'], pd.Timestamp):
                        record['sale_date'] = record['sale_date'].strftime('%Y-%m-%d')
        else:
            return JsonResponse({'status': 'error', 'message': 'Formato de archivo no soportado'})

        # Limpiar datos existentes
        SoldEquipment.objects.all().delete()

        # Insertar nuevos datos
        for item in data:
            SoldEquipment.objects.create(
                serial_number=item.get('serial_number', ''),
                model=item.get('model', ''),
                sale_date=item.get('sale_date', ''),
                customer=item.get('customer', '')
            )

        return JsonResponse({'status': 'success', 'message': 'Datos cargados exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
def upload_service_records(request):
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': 'No se proporcionó ningún archivo'})

        file = request.FILES['file']
        file_extension = file.name.split('.')[-1].lower()

        if file_extension == 'json':
            data = json.loads(file.read().decode('utf-8'))
        elif file_extension in ['xlsx', 'xls']:
            # Leer el archivo Excel
            df = pd.read_excel(file)
            
            # Convertir el DataFrame a una lista de diccionarios
            data = df.to_dict('records')
            
            # Asegurarse de que las fechas estén en el formato correcto
            for record in data:
                if 'service_date' in record:
                    if isinstance(record['service_date'], pd.Timestamp):
                        record['service_date'] = record['service_date'].strftime('%Y-%m-%d')
        else:
            return JsonResponse({'status': 'error', 'message': 'Formato de archivo no soportado'})

        # Limpiar datos existentes
        ServiceRecord.objects.all().delete()

        # Insertar nuevos datos
        for item in data:
            # Convertir la fecha al formato correcto
            service_date = datetime.strptime(item.get('service_date', ''), '%Y-%m-%d').date() if item.get('service_date') else None
            
            # Crear el registro de servicio
            ServiceRecord.objects.create(
                serial_number=item.get('serial_number', ''),
                model=item.get('model', ''),
                service_date=service_date,
                customer=item.get('customer', 'N/A'),
                service_type=item.get('service_type', 'preventive'),
                total_amount=Decimal(str(item.get('total_amount', 0)))
            )

        return JsonResponse({'status': 'success', 'message': 'Datos cargados exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@require_http_methods(["GET"])
def calculate_monthly_pops(request):
    try:
        print("Iniciando cálculo de POPS...")
        
        # Define date range for analysis - desde enero 2023 hasta el mes actual
        start_date = datetime(2023, 1, 1)
        end_date = datetime.now()
        print(f"Rango de fechas: {start_date} a {end_date}")
        
        # Initialize list for monthly results
        monthly_results = []
        
        # Calculate for each month
        current_date = start_date
        while current_date <= end_date:
            try:
                print(f"\nProcesando mes: {current_date.strftime('%Y-%m')}")
                
                # Get the first day of the current month
                month_start = current_date.replace(day=1)
                # Get the last day of the current month
                if current_date.month == 12:
                    month_end = current_date.replace(year=current_date.year + 1, month=1, day=1) - timedelta(days=1)
                else:
                    month_end = current_date.replace(month=current_date.month + 1, day=1) - timedelta(days=1)
                
                print(f"Período del mes: {month_start} a {month_end}")
                
                # Get total equipment sold in the last 10 years up to this month
                ten_years_ago = month_end - timedelta(days=365*10)
                total_equipment = SoldEquipment.objects.filter(
                    sale_date__gte=ten_years_ago,
                    sale_date__lte=month_end
                ).count()
                print(f"Total equipos vendidos (últimos 10 años): {total_equipment}")
                
                # Get service records for the last 12 months up to this month
                twelve_months_ago = month_end - timedelta(days=365)
                
                # Obtener equipos únicos servidos en los últimos 12 meses
                unique_serviced_equipment = ServiceRecord.objects.filter(
                    service_date__gte=twelve_months_ago,
                    service_date__lte=month_end
                ).values('serial_number').distinct()
                
                # Count total unique services
                total_services = unique_serviced_equipment.count()
                print(f"Total equipos únicos servidos en el período: {total_services}")
                
                # Get equipment serial numbers that received service
                serviced_equipment = set(record['serial_number'] for record in unique_serviced_equipment)
                print(f"Equipos con servicio: {len(serviced_equipment)}")
                
                # Count services to recent equipment (sold in last 10 years)
                ten_years_ago = month_end - timedelta(days=365*10)
                recent_equipment = SoldEquipment.objects.filter(
                    sale_date__gte=ten_years_ago,
                    sale_date__lte=month_end
                ).values_list('serial_number', flat=True)
                
                # Count services to recent equipment
                recent_services = sum(1 for sn in serviced_equipment if sn in recent_equipment)
                print(f"Equipos únicos servidos de los últimos 10 años: {recent_services}")
                
                # Count services to older equipment
                older_services = total_services - recent_services
                print(f"Equipos únicos servidos de más de 10 años: {older_services}")
                
                # Calculate revenue for the current month only
                month_revenue = ServiceRecord.objects.filter(
                    service_date__gte=month_start,
                    service_date__lte=month_end
                ).aggregate(
                    total=Sum('total_amount')
                )['total'] or 0
                print(f"Ingresos del mes: ${month_revenue}")
                
                # Calculate POPS for recent equipment (últimos 10 años)
                recent_equipment_count = SoldEquipment.objects.filter(
                    sale_date__gte=ten_years_ago,
                    sale_date__lte=month_end
                ).count()
                print(f"Total equipos vendidos en los últimos 10 años: {recent_equipment_count}")
                
                # POPS para equipos recientes (últimos 10 años)
                if recent_equipment_count > 0:
                    recent_pops = (recent_services / recent_equipment_count) * 100
                else:
                    recent_pops = 0
                print(f"POPS equipos recientes (últimos 10 años): {recent_pops}%")
                
                # POPS para equipos antiguos (más de 10 años)
                # Obtener todos los equipos vendidos antes de los últimos 10 años
                older_equipment = SoldEquipment.objects.filter(
                    sale_date__lt=ten_years_ago
                ).values_list('serial_number', flat=True)
                
                # Contar cuántos de estos equipos antiguos recibieron servicio
                older_serviced_count = sum(1 for sn in serviced_equipment if sn in older_equipment)
                print(f"Equipos antiguos servidos: {older_serviced_count}")
                
                # Calcular POPS para equipos antiguos con respecto al total de equipos vendidos en los últimos 10 años
                if total_equipment > 0:
                    # POPS +10 = Total de equipos únicos servidos / Total de equipos vendidos en los últimos 10 años
                    older_pops = (total_services / total_equipment) * 100
                else:
                    older_pops = 0
                print(f"POPS equipos antiguos (más de 10 años): {older_pops}%")
                
                # Add results for this month
                monthly_results.append({
                    'month': month_start.strftime('%Y-%m'),
                    'total_equipment': total_equipment,
                    'total_services': total_services,
                    'recent_services': recent_services,
                    'older_services': older_services,
                    'total_revenue': float(month_revenue),
                    'recent_pops': round(recent_pops, 2),
                    'older_pops': round(older_pops, 2)
                })
                
                # Move to next month
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
                    
            except Exception as month_error:
                print(f"Error procesando mes {current_date.strftime('%Y-%m')}: {str(month_error)}")
                import traceback
                print(f"Detalles del error: {traceback.format_exc()}")
                # Continue with next month even if there's an error with the current one
                if current_date.month == 12:
                    current_date = current_date.replace(year=current_date.year + 1, month=1)
                else:
                    current_date = current_date.replace(month=current_date.month + 1)
        
        print("\nCalculando top 10 modelos...")
        
        # Calculate top 10 models by service count (sin distinguir por equipos únicos)
        try:
            top_serviced_models = ServiceRecord.objects.values('model').annotate(
                count=Count('id')  # Contar todas las intervenciones, sin distinguir por equipos únicos
            ).order_by('-count')[:10]
            print(f"Top modelos servidos: {list(top_serviced_models)}")
        except Exception as e:
            print(f"Error calculando top modelos servidos: {str(e)}")
            top_serviced_models = []
        
        # Calculate top 10 models by sales
        try:
            top_sold_models = SoldEquipment.objects.values('model').annotate(
                count=Count('serial_number')
            ).order_by('-count')[:10]
            print(f"Top modelos vendidos: {list(top_sold_models)}")
        except Exception as e:
            print(f"Error calculando top modelos vendidos: {str(e)}")
            top_sold_models = []
        
        # Calculate top 10 models by revenue (sin distinguir por equipos únicos)
        try:
            top_revenue_models = ServiceRecord.objects.values('model').annotate(
                total_revenue=Sum('total_amount')
            ).order_by('-total_revenue')[:10]
            
            # Formatear los valores de facturación a 2 decimales
            for model in top_revenue_models:
                if 'total_revenue' in model and model['total_revenue'] is not None:
                    model['total_revenue'] = round(float(model['total_revenue']), 2)
            print(f"Top modelos por facturación: {list(top_revenue_models)}")
        except Exception as e:
            print(f"Error calculando top modelos por facturación: {str(e)}")
            top_revenue_models = []
        
        print("\nPreparando respuesta...")
        response_data = {
            'status': 'success',
            'monthly_results': monthly_results,
            'top_serviced_models': list(top_serviced_models),
            'top_sold_models': list(top_sold_models),
            'top_revenue_models': list(top_revenue_models)
        }
        print(f"Respuesta preparada: {response_data}")
        
        return JsonResponse(response_data)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        print(f"Error calculando POPS: {str(e)}")
        print(f"Detalles del error: {error_details}")
        return JsonResponse({
            'status': 'error',
            'message': f"Error al calcular POPS: {str(e)}"
        })

def export_to_excel(request):
    try:
        # Obtener datos de equipos sin servicios
        equipment_data = get_equipment_without_services_data()
        
        # Crear un nuevo libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Equipos sin Servicios"

        # Definir estilos
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        cell_alignment = Alignment(horizontal='center', vertical='center')

        # Escribir encabezados
        headers = ['Modelo', 'PIN', 'Cliente', 'Fecha de Venta']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Escribir datos
        row = 2
        for model, equipment in equipment_data['equipment_by_model'].items():
            for eq in equipment:
                ws.cell(row=row, column=1, value=model).alignment = cell_alignment
                ws.cell(row=row, column=2, value=eq['serial_number']).alignment = cell_alignment
                ws.cell(row=row, column=3, value=eq['customer']).alignment = cell_alignment
                ws.cell(row=row, column=4, value=eq['sale_date']).alignment = cell_alignment
                row += 1

        # Ajustar anchos de columna
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Crear respuesta
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=equipos_sin_servicios.xlsx'
        
        # Guardar el libro
        wb.save(response)
        
        return response
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)})

@csrf_exempt
def get_equipment_without_services(request):
    try:
        # Obtener fechas de filtro de la solicitud
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        
        # Si no se proporcionan fechas, usar el rango completo de registros de servicio
        if not start_date or not end_date:
            service_dates = ServiceRecord.objects.aggregate(
                earliest_date=Min('service_date'),
                latest_date=Max('service_date')
            )
            start_date = service_dates['earliest_date']
            end_date = service_dates['latest_date']
        else:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # Obtener equipos vendidos en todo el período
        sold_equipment = SoldEquipment.objects.all().values('serial_number', 'model', 'sale_date', 'customer')
        
        # Obtener servicios en el rango de fechas
        service_records = ServiceRecord.objects.filter(
            service_date__gte=start_date,
            service_date__lte=end_date
        ).values('serial_number')
        
        # Obtener serial numbers que tienen servicios
        serviced_serials = set(record['serial_number'] for record in service_records)
        
        # Filtrar equipos sin servicios y agrupar por modelo
        equipment_by_model = {}
        for record in sold_equipment:
            try:
                # Asegurarse de que sale_date sea un objeto date
                sale_date = record['sale_date']
                if isinstance(sale_date, str):
                    sale_date = datetime.strptime(sale_date, '%Y-%m-%d').date()
                elif isinstance(sale_date, datetime):
                    sale_date = sale_date.date()
                
                # Solo agregar si no tiene servicios
                if record['serial_number'] not in serviced_serials:
                    model = str(record['model']) if record['model'] else 'Sin Modelo'
                    if model not in equipment_by_model:
                        equipment_by_model[model] = []
                    
                    equipment_data = {
                        'serial_number': str(record['serial_number']) if record['serial_number'] else '',
                        'model': model,
                        'customer': str(record['customer']) if record['customer'] is not None else 'N/A',
                        'sale_date': sale_date.strftime('%Y-%m-%d')
                    }
                    equipment_by_model[model].append(equipment_data)
            except Exception as e:
                print(f"Error processing record {record}: {str(e)}")
                continue
        
        # Convertir el diccionario agrupado en una lista plana para mantener compatibilidad
        equipment_without_services = []
        for model, equipment_list in equipment_by_model.items():
            equipment_without_services.extend(equipment_list)
        
        # Obtener el rango completo de fechas de venta
        sale_dates = []
        for record in sold_equipment:
            try:
                sale_date = record['sale_date']
                if isinstance(sale_date, str):
                    sale_date = datetime.strptime(sale_date, '%Y-%m-%d').date()
                elif isinstance(sale_date, datetime):
                    sale_date = sale_date.date()
                sale_dates.append(sale_date)
            except Exception as e:
                print(f"Error processing sale date for record {record}: {str(e)}")
                continue
        
        earliest_date = min(sale_dates) if sale_dates else datetime.now().date()
        latest_date = max(sale_dates) if sale_dates else datetime.now().date()
        
        # Calcular totales por modelo
        model_totals = {model: len(equipment) for model, equipment in equipment_by_model.items()}
        
        return JsonResponse({
            'status': 'success',
            'equipment_without_services': equipment_without_services,
            'equipment_by_model': equipment_by_model,
            'model_totals': model_totals,
            'total_sold': len(sold_equipment),
            'total_without_service': len(equipment_without_services),
            'sales_period': f"{earliest_date.strftime('%Y-%m-%d')} hasta {latest_date.strftime('%Y-%m-%d')}",
            'services_period': f"{start_date.strftime('%Y-%m-%d')} hasta {end_date.strftime('%Y-%m-%d')}"
        })
        
    except Exception as e:
        print(f"Error in get_equipment_without_services: {str(e)}")
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })
